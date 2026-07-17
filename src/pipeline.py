"""
pipeline.py — Data extraction pipeline adapted from Colab notebook.

Converts the sequential Colab workflow into reusable functions:
    1. preprocess_image   — CLAHE + sharpen + resize (OpenCV)
    2. extract_from_zip   — Extract ZIP → list of (filename, bytes)
    3. extract_packages_gemini — Gemini OCR with retry + key rotation
    4. clean_dataframe    — Normalize provider, compute yield & category
    5. generate_excel     — openpyxl horizontal layout + heatmap
"""

from __future__ import annotations

import io
import json
import math
import time
import zipfile
from pathlib import Path

import cv2
import numpy as np
import pandas as pd
from PIL import Image

# ══════════════════════════════════════════════════════════════════════════════
# Constants
# ══════════════════════════════════════════════════════════════════════════════

PROVIDER_CODES = ["IM3", "3ID", "TSEL", "XL", "AXIS", "SF"]

PROVIDER_MAPPING = {
    "INDOSAT": "IM3",
    "THREE": "3ID",
    "TELKOMSEL": "TSEL",
    "SMARTFREN": "SF",
    "XL": "XL",
    "AXIS": "AXIS",
    "3": "3ID",
}

CATEGORIES_ORDER = [
    "SACHET 1D-2D",
    "SACHET 3D",
    "SACHET 5D",
    "SACHET 7D",
    "SACHET 10D-15D",
    "MONTHLY 30-50K",
    "MONTHLY > 50K",
    "OTHER",
]

HEADER_COLORS = {
    "IM3": "FFC000",
    "3ID": "FF00FF",
    "TSEL": "FF0000",
    "XL": "0070C0",
    "AXIS": "7030A0",
    "SF": "FF00FF",
}

# Konfigurasi Model
MODEL_NAME = "gemini-3.1-flash-lite"

EXTRACTION_PROMPT = """
Ekstrak semua paket data internet dari gambar ini.

ATURAN:
- Provider harus salah satu dari: "TSEL", "IM3", "3ID", "XL", "AXIS", "SF"
- Konversi harga: K = ribuan (37K -> 37000). Abaikan "Rp", titik, spasi.
- UNLIMITED + GB/hari: kalikan (contoh 3GB/hari x 28hari = 84.0 GB)
- price: integer, gb: float, days: integer
- Abaikan baris jika price, gb, atau days kosong
- Abaikan watermark, botol, rak, orang, dan objek non-data
- Ekstrak juga `image_timestamp` (waktu pengambilan foto) dan `image_location` (lokasi/geolokasi) jika ada di dalam gambar overlay. Jika tidak ada, isi null.

Kembalikan JSON array saja, tanpa teks tambahan:
[{"provider":"...","price":0,"gb":0.0,"days":0,"image_timestamp":"2024-01-01 12:00","image_location":"Jakarta"}]
""".strip()

_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tiff"}


# ══════════════════════════════════════════════════════════════════════════════
# 1. Image Preprocessing (from notebook cells 55-109)
# ══════════════════════════════════════════════════════════════════════════════

def preprocess_image(image_bytes: bytes) -> bytes:
    """Apply CLAHE contrast enhancement + sharpening + resize (max 1800px width).

    Parameters
    ----------
    image_bytes : bytes
        Raw image file bytes (any format OpenCV can read).

    Returns
    -------
    bytes
        JPEG-encoded preprocessed image bytes.
    """
    # Decode from bytes
    arr = np.frombuffer(image_bytes, dtype=np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if img is None:
        raise ValueError("Gagal membaca gambar. Format tidak didukung.")

    h, w = img.shape[:2]

    # Resize if wider than 1800px
    if w > 1800:
        scale = 1800 / w
        img = cv2.resize(img, (1800, int(h * scale)))

    # CLAHE on L channel (LAB color space)
    lab = cv2.cvtColor(img, cv2.COLOR_BGR2LAB)
    l_ch, a_ch, b_ch = cv2.split(lab)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    l_ch = clahe.apply(l_ch)
    lab = cv2.merge((l_ch, a_ch, b_ch))
    img = cv2.cvtColor(lab, cv2.COLOR_LAB2BGR)

    # Sharpen
    kernel = np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]])
    img = cv2.filter2D(img, -1, kernel)

    # Encode back to JPEG bytes
    success, encoded = cv2.imencode(".jpg", img, [cv2.IMWRITE_JPEG_QUALITY, 95])
    if not success:
        raise ValueError("Gagal meng-encode gambar hasil preprocessing.")
    return encoded.tobytes()


# ══════════════════════════════════════════════════════════════════════════════
# 2. ZIP Extraction
# ══════════════════════════════════════════════════════════════════════════════

def is_image_file(filename: str) -> bool:
    """Check if filename has an image extension."""
    return Path(filename).suffix.lower() in _IMAGE_EXTENSIONS


def is_zip_file(filename: str, data: bytes | None = None) -> bool:
    """Check if a file is a ZIP archive (by extension or magic bytes)."""
    if Path(filename).suffix.lower() == ".zip":
        return True
    if data and len(data) >= 4 and data[:4] == b"PK\x03\x04":
        return True
    return False


def extract_from_zip(zip_bytes: bytes) -> list[tuple[str, bytes]]:
    """Extract image files from a ZIP archive.

    Parameters
    ----------
    zip_bytes : bytes
        Raw ZIP file bytes.

    Returns
    -------
    list[tuple[str, bytes]]
        List of (filename, image_bytes) for each image found.
    """
    results = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes), "r") as zf:
        for name in sorted(zf.namelist()):
            # Skip directories and hidden/system files
            if name.endswith("/") or name.startswith("__MACOSX"):
                continue
            if is_image_file(name):
                data = zf.read(name)
                results.append((Path(name).name, data))
    return results


def _parse_gemini_response(raw_text: str) -> list[dict]:
    """Parse JSON from Gemini response. Handles both structured JSON output
    and fallback text responses with embedded JSON."""
    import re
    import json
    text = raw_text.strip()
    
    # Strategy 1: Direct parse (works with response_mime_type="application/json")
    try:
        data = json.loads(text)
        if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
            return data
    except json.JSONDecodeError:
        pass
    
    # Strategy 2: Strip markdown fences
    cleaned = text.replace("```json", "").replace("```", "").strip()
    try:
        data = json.loads(cleaned)
        if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
            return data
    except json.JSONDecodeError:
        pass
    
    # Strategy 3: Find JSON array pattern [{...}] in text
    match = re.search(r'\[\s*\{[\s\S]*\}\s*\]', cleaned)
    if match:
        try:
            data = json.loads(match.group(0))
            if isinstance(data, list) and len(data) > 0:
                return data
        except json.JSONDecodeError:
            pass
    
    raise RuntimeError(
        f"Gagal mem-parse respons Gemini sebagai JSON.\n"
        f"Respons mentah:\n{text[:500]}"
    )


def extract_packages_gemini(
    image_bytes_list: list[bytes],
    api_keys: list[str],
    key_index: int = 0,
    max_retries: int = 6,
    model: str = MODEL_NAME,
    on_status: callable | None = None,
    custom_prompt: str | None = None,
) -> tuple[list[dict], int]:
    """Extract packages from multiple preprocessed images via Gemini.

    Parameters
    ----------
    image_bytes_list : list[bytes]
        List of preprocessed JPEG image bytes.
    api_keys : list[str]
        List of Gemini API keys for rotation.
    key_index : int
        Starting index in api_keys (for rotation across images).
    max_retries : int
        Max retry attempts per image.
    model : str
        Gemini model name.
    on_status : callable | None
        Optional callback ``on_status(message: str)`` for status updates.
    custom_prompt : str | None
        Optional user prompt to append to the system prompt.

    Returns
    -------
    tuple[list[dict], int]
        (extracted_packages, updated_key_index)
    """
    from google import genai
    from google.genai import types

    current_key_idx = key_index % len(api_keys)

    for attempt in range(max_retries):
        try:
            client = genai.Client(api_key=api_keys[current_key_idx])

            # Build final prompt
            final_prompt = EXTRACTION_PROMPT
            if custom_prompt and custom_prompt.strip() and custom_prompt.strip() != "Tolong scan gambar ini.":
                final_prompt = f"{final_prompt}\n\nINSTRUKSI TAMBAHAN DARI USER:\n{custom_prompt.strip()}"

            # Build image parts from all bytes
            parts = [final_prompt]
            for img_bytes in image_bytes_list:
                parts.append(types.Part.from_bytes(
                    data=img_bytes, mime_type="image/jpeg"
                ))

            # Use structured JSON output for speed (no persona logs)
            config = types.GenerateContentConfig(
                response_mime_type="application/json",
            )

            response = client.models.generate_content(
                model=model,
                contents=parts,
                config=config,
            )

            data = _parse_gemini_response(response.text)
            return data, current_key_idx

        except Exception as e:
            error_msg = str(e)
            if "503" in error_msg or "UNAVAILABLE" in error_msg or "429" in error_msg:
                # Rotate to next API key
                current_key_idx = (current_key_idx + 1) % len(api_keys)
                msg = (
                    f"API Limit/Server Sibuk (Percobaan {attempt + 1}/{max_retries}). "
                    f"Mencoba API Key berikutnya..."
                )
                if on_status:
                    on_status(msg)
                time.sleep(2)
            else:
                raise RuntimeError(f"Gagal ekstraksi Gemini: {error_msg}")

    raise RuntimeError(
        f"Gagal setelah {max_retries} percobaan. Semua API key telah dicoba."
    )


# ══════════════════════════════════════════════════════════════════════════════
# 4. Post-Processing (from notebook cells 280-432)
# ══════════════════════════════════════════════════════════════════════════════

def compute_yield(price: float, gb: float) -> int:
    """Compute yield = ceil(price / gb). Returns 0 if gb is zero or inf."""
    if gb <= 0:
        return 0
    val = price / gb
    if val == np.inf or np.isnan(val):
        return 0
    return math.ceil(val)


def categorize(days: int, price: int) -> str:
    """Categorize package based on days and price (notebook logic)."""
    if days <= 2:
        return "SACHET 1D-2D"
    elif days == 3:
        return "SACHET 3D"
    elif days <= 5:
        return "SACHET 5D"
    elif days <= 7:
        return "SACHET 7D"
    elif days <= 15:
        return "SACHET 10D-15D"
    elif days >= 28 and price <= 50000:
        return "MONTHLY 30-50K"
    elif days >= 28 and price > 50000:
        return "MONTHLY > 50K"
    else:
        return "OTHER"


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize provider names, compute yield & category.

    Parameters
    ----------
    df : pd.DataFrame
        Raw extraction result with columns: provider, price, gb, days.

    Returns
    -------
    pd.DataFrame
        Cleaned DataFrame with added yield_val and category columns.
    """
    df = df.copy()

    # Normalize provider codes
    df["provider"] = (
        df["provider"]
        .astype(str)
        .str.strip()
        .str.upper()
        .map(PROVIDER_MAPPING)
        .fillna(df["provider"].astype(str).str.strip().str.upper())
    )

    # Ensure numeric types
    df["price"] = pd.to_numeric(df["price"], errors="coerce")
    df["gb"] = pd.to_numeric(df["gb"], errors="coerce")
    df["days"] = pd.to_numeric(df["days"], errors="coerce")

    # Drop rows with missing critical data
    df = df.dropna(subset=["price", "gb", "days"])

    # Cast to proper types
    df["price"] = df["price"].astype(int)
    df["days"] = df["days"].astype(int)

    # Compute yield and category
    df["yield_val"] = df.apply(
        lambda row: compute_yield(row["price"], row["gb"]), axis=1
    )
    df["category"] = df.apply(
        lambda row: categorize(row["days"], row["price"]), axis=1
    )

    # Sort by category → provider → price
    df = df.sort_values(by=["category", "provider", "price"]).reset_index(drop=True)

    return df


# ══════════════════════════════════════════════════════════════════════════════
# 4.5 Summary Insights Generator
# ══════════════════════════════════════════════════════════════════════════════

def generate_summary_insights(df: pd.DataFrame) -> dict[str, list[str]]:
    """Generate best offer insights grouped by category.
    
    Returns a dict with two keys:
    - 'sachet': list of best offers for SACHET categories
    - 'monthly': list of best offers for MONTHLY categories
    """
    insights = {
        "sachet": [],
        "monthly": [],
    }
    
    # Category mappings for display
    category_labels = {
        "SACHET 1D-2D": "1-2 days",
        "SACHET 3D": "3 days",
        "SACHET 5D": "5 days",
        "SACHET 7D": "7 days",
        "SACHET 10D-15D": ">10 days",
        "MONTHLY 30-50K": "30K-50K",
        "MONTHLY > 50K": ">100K",
    }
    
    # Also add 50-100K if we have it
    price_ranges = {
        "MONTHLY 30-50K": "30K-50K",
        "MONTHLY 50K-100K": "50K-100K",
        "MONTHLY > 50K": ">100K",
    }
    
    # Process SACHET categories
    for cat in ["SACHET 1D-2D", "SACHET 3D", "SACHET 5D", "SACHET 7D", "SACHET 10D-15D"]:
        cat_df = df[df["category"] == cat]
        if not cat_df.empty:
            # Find best deal (lowest yield)
            best_idx = cat_df["yield_val"].idxmin()
            best = cat_df.loc[best_idx]
            
            label = category_labels.get(cat, cat)
            insight = f"{label:15} : {best['gb']:.0f}GB {best['price']:,} ({best['provider']})"
            insights["sachet"].append(insight)
    
    # Process MONTHLY categories with a special case for 50-100K range
    # Check if we have data in that range
    monthly_df = df[df["category"].str.startswith("MONTHLY", na=False)]
    
    # Split by price range
    monthly_30_50k = monthly_df[(monthly_df["price"] >= 30000) & (monthly_df["price"] <= 50000)]
    monthly_50_100k = monthly_df[(monthly_df["price"] > 50000) & (monthly_df["price"] <= 100000)]
    monthly_over_100k = monthly_df[monthly_df["price"] > 100000]
    
    for range_df, label in [
        (monthly_30_50k, "30K-50K"),
        (monthly_50_100k, "50K-100K"),
        (monthly_over_100k, ">100K"),
    ]:
        if not range_df.empty:
            best_idx = range_df["yield_val"].idxmin()
            best = range_df.loc[best_idx]
            insight = f"{label:15} : {best['gb']:.0f}GB {best['price']:,} ({best['provider']})"
            insights["monthly"].append(insight)
    
    return insights


# ══════════════════════════════════════════════════════════════════════════════
# 5. Excel Generation (from notebook cells 389-546)
# ══════════════════════════════════════════════════════════════════════════════

def generate_excel(df: pd.DataFrame) -> bytes:
    """Generate the formatted Excel report matching the Colab notebook layout.

    The layout is horizontal: each provider gets 4 columns (Price, GB, Days, Yield),
    grouped by category rows. Best-deal rows get red dashed borders.
    Yield cells get a green-yellow-red conditional formatting heatmap.

    Parameters
    ----------
    df : pd.DataFrame
        Cleaned DataFrame with columns: provider, price, gb, days, yield_val, category.

    Returns
    -------
    bytes
        The .xlsx file content as bytes.
    """
    from openpyxl import Workbook
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Harga Automasi"

    # ── Styles ────────────────────────────────────────────────────────────
    font_bold_white = Font(bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    border_thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    border_red_dashed = Border(
        left=Side(style="mediumDashed", color="FF0000"),
        right=Side(style="mediumDashed", color="FF0000"),
        top=Side(style="mediumDashed", color="FF0000"),
        bottom=Side(style="mediumDashed", color="FF0000"),
    )

    providers_order = PROVIDER_CODES

    # ── Generate Header Row 1 (Provider names, merged) ────────────────────
    col_start = 2
    for prov in providers_order:
        ws.merge_cells(
            start_row=1, start_column=col_start,
            end_row=1, end_column=col_start + 3,
        )
        c = ws.cell(row=1, column=col_start, value=prov)
        c.fill = PatternFill(
            start_color=HEADER_COLORS.get(prov, "808080"), fill_type="solid"
        )
        c.font = font_bold_white
        c.alignment = align_center
        c.border = border_thin

        # Sub-headers (row 2): Price, GB, Days, Yield
        for i, sh in enumerate(["Price", "GB", "Days", "Yield"]):
            c_sub = ws.cell(row=2, column=col_start + i, value=sh)
            c_sub.fill = (
                PatternFill(start_color="FFD966", fill_type="solid")
                if sh == "Yield"
                else PatternFill(start_color="D9D9D9", fill_type="solid")
            )
            c_sub.font = Font(bold=True)
            c_sub.alignment = align_center
            c_sub.border = border_thin

        col_start += 4

    # ── Generate Body Data (Horizontal Layout) ────────────────────────────
    current_row = 3
    yield_cells: list[str] = []

    for cat in CATEGORIES_ORDER:
        cat_df = df[df["category"] == cat]
        if cat_df.empty:
            continue

        # Group data per provider
        prov_data = {
            prov: cat_df[cat_df["provider"] == prov].to_dict("records")
            for prov in providers_order
        }
        max_rows = max([len(v) for v in prov_data.values()] + [1])

        # Category label (merged vertically)
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row + max_rows - 1, end_column=1,
        )
        c_cat = ws.cell(row=current_row, column=1, value=cat)
        c_cat.alignment = Alignment(
            vertical="center", horizontal="center", wrap_text=True
        )
        c_cat.border = border_thin

        # Fill rows
        for i in range(max_rows):
            col_idx = 2

            # Find minimum yield in this row for "best deal" highlight
            row_yields = []
            for prov in providers_order:
                if i < len(prov_data[prov]):
                    row_yields.append(prov_data[prov][i]["yield_val"])
            min_yield_in_row = min(row_yields) if row_yields else None

            for prov in providers_order:
                # Default border for all cells (even empty ones)
                for offset in range(4):
                    ws.cell(
                        row=current_row + i, column=col_idx + offset
                    ).border = border_thin

                if i < len(prov_data[prov]):
                    pkg = prov_data[prov][i]

                    ws.cell(
                        row=current_row + i, column=col_idx, value=pkg["price"]
                    ).number_format = "#,##0"
                    ws.cell(
                        row=current_row + i, column=col_idx + 1, value=pkg["gb"]
                    )
                    ws.cell(
                        row=current_row + i, column=col_idx + 2, value=pkg["days"]
                    )

                    c_yield = ws.cell(
                        row=current_row + i,
                        column=col_idx + 3,
                        value=pkg["yield_val"],
                    )
                    c_yield.number_format = "#,##0"
                    yield_cells.append(c_yield.coordinate)

                    # Highlight best deal (red dashed border)
                    if (
                        min_yield_in_row is not None
                        and pkg["yield_val"] == min_yield_in_row
                        and pkg["yield_val"] > 0
                    ):
                        for offset in range(4):
                            ws.cell(
                                row=current_row + i, column=col_idx + offset
                            ).border = border_red_dashed

                col_idx += 4

        current_row += max_rows

    # ── Conditional Formatting (Yield Heatmap) ────────────────────────────
    color_scale = ColorScaleRule(
        start_type="min",
        start_color="63BE7B",
        mid_type="percentile",
        mid_value=50,
        mid_color="FFEB84",
        end_type="max",
        end_color="F8696B",
    )
    if yield_cells:
        ws.conditional_formatting.add(f"E3:Z{current_row}", color_scale)

    # ── Auto-adjust column width ──────────────────────────────────────────
    for col_idx_w in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx_w)
        max_length = 0
        for cell in ws[column_letter]:
            if type(cell).__name__ == "MergedCell":
                continue
            if cell.value:
                try:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except Exception:
                    pass
        ws.column_dimensions[column_letter].width = max_length + 2

    ws.column_dimensions["A"].width = 15  # Fix category column width

    # ── Generate Summary Insights (Notes Section) ─────────────────────────
    insights = generate_summary_insights(df)
    
    # Add Notes section starting from a few rows below the data
    notes_start_row = current_row + 3
    
    # Title
    c_title = ws.cell(row=notes_start_row, column=1, value="NOTES:")
    c_title.font = Font(bold=True, size=12, color="000000")
    
    # Subtitle 1: Best offer based on validity
    notes_row = notes_start_row + 2
    c_subtitle = ws.cell(row=notes_row, column=1, value="The best offer based on validity:")
    c_subtitle.font = Font(bold=True, size=10)
    
    # List SACHET insights
    notes_row += 1
    for insight in insights["sachet"]:
        ws.cell(row=notes_row, column=1, value=f"• {insight}")
        notes_row += 1
    
    # Subtitle 2: Best offer monthly pack
    notes_row += 1
    c_subtitle2 = ws.cell(row=notes_row, column=1, value="The best offer monthly pack:")
    c_subtitle2.font = Font(bold=True, size=10)
    
    # List MONTHLY insights
    notes_row += 1
    for insight in insights["monthly"]:
        ws.cell(row=notes_row, column=1, value=f"• {insight}")
        notes_row += 1

    # ── Save to bytes ─────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
